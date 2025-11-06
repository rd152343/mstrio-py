"""
Zebra Objects Lister - Clean Version

This script lists all objects within a specified MicroStrategy project and exports to Excel.
Uses location attribute for complete folder paths.

Author: Zebra Technologies
Date: November 2025
"""

import sys
import os
import pandas as pd
from datetime import datetime
import time

# Add the workflows directory to path to import other modules
sys.path.append(os.path.dirname(__file__))

try:
    from mstrio.object_management import list_objects
    from mstrio.server import Project
    from mstrio.types import ObjectTypes
    from mstrio.object_management.search_enums import SearchDomain
    print("✓ Successfully imported mstrio modules")
except ImportError as e:
    print(f"✗ Error importing mstrio modules: {e}")
    sys.exit(1)

# Import connection modules
try:
    import ZebraCreateConnection
    import ZebraCloseConnection
    print("✓ Successfully imported Zebra connection modules")
except ImportError as e:
    print(f"✗ Error importing Zebra connection modules: {e}")
    print("Make sure ZebraCreateConnection.py and ZebraCloseConnection.py are in the same directory")
    sys.exit(1)


def get_zebra_connection():
    """Get connection using ZebraCreateConnection."""
    try:
        print("Creating connection...")
        conn = ZebraCreateConnection.main()
        if conn:
            print("✓ Connection established successfully")
            return conn
        else:
            print("✗ Failed to create connection")
            return None
    except Exception as e:
        print(f"✗ Error creating connection: {e}")
        return None


def close_zebra_connection(conn):
    """Close the connection."""
    try:
        if conn:
            print("\nClosing connection...")
            conn.close()
            print("✓ Connection closed successfully")
    except Exception as e:
        print(f"✗ Error closing connection: {e}")


def get_object_absolute_path(obj):
    """Get the complete absolute path for an object using the location attribute."""
    try:
        obj_name = getattr(obj, 'name', 'Unknown')
        
        # Use the 'location' attribute which contains the full path
        location = getattr(obj, 'location', None)
        if location and isinstance(location, str) and location.strip():
            full_path = location.strip()
            
            # If location includes object name at the end, use as-is
            # Otherwise append object name
            if full_path.endswith(obj_name):
                return full_path
            else:
                return f"{full_path}/{obj_name}"
        
        # Fallback to root level if no location
        return f"/{obj_name}"
        
    except Exception:
        obj_name = getattr(obj, 'name', 'Unknown')
        return f"/{obj_name}"


def list_all_objects(conn, project_id):
    """List all objects in the specified project with complete absolute paths."""
    all_objects = []
    project_name = "Unknown"
    
    try:
        start_time = time.time()
        print(f"\n=== Listing All Objects in Project ===")
        print(f"Project ID: {project_id}")
        
        # Connect to project
        try:
            project = Project(connection=conn, id=project_id)
            project_name = project.name
            print(f"✓ Connected to project: {project_name}")
        except Exception as e:
            print(f"✗ Failed to connect to project: {e}")
            return all_objects, project_name
        
        # Define object types to search for
        object_types_to_search = [
            ('Attributes', ObjectTypes.ATTRIBUTE),            
            ('Metrics', ObjectTypes.METRIC)
            # ,            
            # ('Filters', ObjectTypes.FILTER),
            # ('Folders', ObjectTypes.FOLDER),
            # ('Facts', ObjectTypes.FACT),            
            # ('Tables', ObjectTypes.TABLE),                                    
            # ('Prompts', ObjectTypes.PROMPT),
            # ('Security Filters', ObjectTypes.SECURITY_FILTER),
            # ('Reports', ObjectTypes.REPORT_DEFINITION),
            # ('Documents/Dashboards', ObjectTypes.DOCUMENT_DEFINITION), 
            # ('Dashboard Views', ObjectTypes.DASHBOARD_PERSONAL_VIEW),
            # ('Templates', ObjectTypes.TEMPLATE)

            #('AttributeForms', ObjectTypes.ATTRIBUTE_FORM)
        ]
        
        print("🔍 Scanning for objects...")
        total_found = 0
        
        for type_name, object_type in object_types_to_search:
            try:
                print(f"  Searching {type_name}...")
                
                objects = list_objects(
                    connection=conn, 
                    object_type=object_type,
                    project_id=project_id,
                    domain=SearchDomain.PROJECT
                )
                
                if objects:
                    print(f"    Found {len(objects)} {type_name}")
                    
                    # Process objects
                    processed_count = 0
                    error_count = 0
                    
                    for obj in objects:
                        try:
                            obj_id = getattr(obj, 'id', 'N/A')
                            obj_name = getattr(obj, 'name', 'N/A')
                            
                            # Get absolute path using location attribute
                            absolute_path = get_object_absolute_path(obj)
                            
                            # Extract parent folder from absolute path
                            if '/' in absolute_path:
                                path_parts = absolute_path.rstrip('/').split('/')
                                if len(path_parts) > 1:
                                    if path_parts[-1] == obj_name:
                                        parent_folder = '/'.join(path_parts[:-1])
                                    else:
                                        parent_folder = absolute_path
                                    
                                    if not parent_folder.startswith('/'):
                                        parent_folder = '/' + parent_folder
                                else:
                                    parent_folder = '/'
                            else:
                                parent_folder = '/'
                            
                            # Extract first folder name from path (after project name)
                            first_folder = 'Root'  # Default if no folders
                            if '/' in absolute_path:
                                path_parts = absolute_path.strip('/').split('/')
                                # Filter out empty parts and project name if present
                                filtered_parts = [part for part in path_parts if part and part != project_name]
                                if filtered_parts:
                                    # Get the first folder after filtering
                                    first_folder = filtered_parts[0]
                                else:
                                    first_folder = 'Root'
                            
                            # Get description
                            try:
                                obj_desc = getattr(obj, 'description', 'N/A')
                                if obj_desc and len(obj_desc) > 100:
                                    obj_desc = obj_desc[:100] + '...'
                            except Exception:
                                obj_desc = 'N/A'
                            
                            obj_info = {
                                'id': obj_id,
                                'name': obj_name,
                                'absolute_path': absolute_path,
                                'parent_folder': parent_folder,
                                'first_folder': first_folder,
                                'type': type_name,
                                'subtype': getattr(obj, 'subtype', 'N/A'),
                                'description': obj_desc,
                                'status': 'accessible'
                            }
                            all_objects.append(obj_info)
                            processed_count += 1
                            
                        except Exception as e:
                            error_count += 1
                            if "ERR004" in str(e) or "does not exist in the metadata" in str(e):
                                # Handle orphaned objects
                                try:
                                    obj_id = getattr(obj, 'id', 'N/A')
                                    orphaned_name = f"ORPHANED_OBJECT_{error_count}"
                                    
                                    obj_info = {
                                        'id': obj_id,
                                        'name': orphaned_name,
                                        'absolute_path': f"/ORPHANED/{orphaned_name}",
                                        'parent_folder': '/ORPHANED',
                                        'first_folder': 'ORPHANED',
                                        'type': type_name,
                                        'subtype': 'ORPHANED',
                                        'description': f'Metadata error: {str(e)[:50]}...',
                                        'status': 'orphaned'
                                    }
                                    all_objects.append(obj_info)
                                except:
                                    pass
                            continue
                    
                    if error_count > 0:
                        print(f"    Processed: {processed_count}, Errors: {error_count}")
                    
                    total_found += len(objects)
                else:
                    print(f"    No {type_name} found")
                    
            except Exception as e:
                print(f"    Warning: Error searching {type_name}: {e}")
                continue
        
        total_time = time.time() - start_time
        print(f"\n✓ Search completed! Found {total_found} total objects")
        print(f"📊 Total processing time: {total_time:.2f} seconds")
        
    except Exception as e:
        print(f"✗ Error in list_all_objects: {e}")
    
    return all_objects, project_name


def export_objects_to_excel(objects, project_id, project_name="Unknown"):
    """Export objects list to Excel with formatting."""
    if not objects:
        print("⚠ No objects to export")
        return None
    
    try:
        # Create DataFrame
        df = pd.DataFrame(objects)
        
        # Reorder columns for better readability
        desired_column_order = ['id', 'name', 'type', 'subtype', 'first_folder', 'parent_folder', 'absolute_path', 'description', 'status']
        
        # Ensure all desired columns exist, add any missing columns that might be in the data
        existing_columns = df.columns.tolist()
        reordered_columns = []
        
        # Add desired columns in order if they exist
        for col in desired_column_order:
            if col in existing_columns:
                reordered_columns.append(col)
        
        # Add any remaining columns that weren't in the desired order
        for col in existing_columns:
            if col not in reordered_columns:
                reordered_columns.append(col)
        
        # Reorder the dataframe
        df = df[reordered_columns]
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        clean_project_name = "".join(c for c in project_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        if not clean_project_name or clean_project_name == "Unknown":
            filename = f"MicroStrategy_Objects_{project_id}_{timestamp}.xlsx"
        else:
            filename = f"MicroStrategy_Objects_{clean_project_name}_{timestamp}.xlsx"
        
        print(f"\n📊 Exporting {len(objects)} objects to Excel...")
        print(f"   Filename: {filename}")
        
        # Create Excel writer with formatting options
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Write main data
            df.to_excel(writer, sheet_name='Objects', index=False, startrow=1)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Objects']
            
            # Add header information
            worksheet['A1'] = f"MicroStrategy Objects Report - Project: {project_name} (ID: {project_id})"
            
            # Format header row
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Header styling
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Title styling
            title_font = Font(bold=True, size=14, color="1F4E79")
            title_alignment = Alignment(horizontal="left", vertical="center")
            
            # Apply title formatting
            worksheet['A1'].font = title_font
            worksheet['A1'].alignment = title_alignment
            
            # Apply header formatting
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=2, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Auto-adjust column widths with special handling for path columns
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                column_name = column[1].value if len(column) > 1 else ""
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set column-specific width limits
                if column_name in ['absolute_path']:
                    adjusted_width = min(max_length + 2, 80)
                    adjusted_width = max(adjusted_width, 25)
                elif column_name in ['parent_folder']:
                    adjusted_width = min(max_length + 2, 60)
                    adjusted_width = max(adjusted_width, 20)
                elif column_name in ['first_folder']:
                    adjusted_width = min(max_length + 2, 25)
                    adjusted_width = max(adjusted_width, 15)
                elif column_name in ['description']:
                    adjusted_width = min(max_length + 2, 40)
                    adjusted_width = max(adjusted_width, 15)
                elif column_name in ['name']:
                    adjusted_width = min(max_length + 2, 35)
                    adjusted_width = max(adjusted_width, 15)
                else:
                    adjusted_width = min(max_length + 2, 25)
                    adjusted_width = max(adjusted_width, 10)
                
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add borders to data
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 2, min_col=1, max_col=len(df.columns)):
                for cell in row:
                    cell.border = thin_border
            
            # Color code orphaned objects
            if 'status' in df.columns:
                status_col_idx = df.columns.get_loc('status') + 1
                orphaned_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                
                for row_idx, status in enumerate(df['status'], start=3):
                    if status == 'orphaned':
                        for col in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row_idx, column=col).fill = orphaned_fill
        
        print(f"✓ Excel export completed successfully!")
        print(f"   File saved as: {filename}")
        
        # Create summary statistics
        print(f"\n📈 Export Summary:")
        print(f"   Total objects exported: {len(objects)}")
        
        # Type breakdown
        type_counts = df['type'].value_counts()
        print(f"   Object types:")
        for obj_type, count in type_counts.items():
            print(f"     - {obj_type}: {count}")
        
        # Status breakdown
        if 'status' in df.columns:
            status_counts = df['status'].value_counts()
            print(f"   Object status:")
            for status, count in status_counts.items():
                print(f"     - {status.title()}: {count}")
        
        # Path information
        if 'absolute_path' in df.columns:
            unique_folders = df['parent_folder'].nunique() if 'parent_folder' in df.columns else 0
            print(f"   Path information:")
            print(f"     - Objects with paths: {len(df[~df['absolute_path'].str.startswith('/')])}")
            print(f"     - Unique parent folders: {unique_folders}")
        
        return filename
        
    except Exception as e:
        print(f"✗ Error exporting to Excel: {e}")
        return None


def main():
    """Main function to list all objects in a project and export to Excel."""
    print("=== Zebra Objects Lister ===")
    
    # Get connection
    conn = get_zebra_connection()
    if not conn:
        print("✗ Failed to establish connection. Exiting.")
        return
    
    # Get project ID from user
    try:
        project_id = input("\nEnter Project ID: ").strip()
        if not project_id:
            print("✗ Project ID is required!")
            return
    except KeyboardInterrupt:
        print("\n✗ Operation cancelled by user")
        return
    except Exception as e:
        print(f"✗ Error getting project ID: {e}")
        return
    
    try:
        # List all objects
        objects, project_name = list_all_objects(conn, project_id)
        
        if objects:
            print(f"\n=== Summary ===")
            print(f"Total objects found: {len(objects)}")
            
            # Group by type and status
            type_counts = {}
            status_counts = {'accessible': 0, 'orphaned': 0}
            
            for obj in objects:
                obj_type = obj.get('type', 'Unknown')
                obj_status = obj.get('status', 'accessible')
                
                type_counts[obj_type] = type_counts.get(obj_type, 0) + 1
                status_counts[obj_status] = status_counts.get(obj_status, 0) + 1
            
            print("\nObjects by type:")
            for obj_type, count in sorted(type_counts.items()):
                print(f"  {obj_type}: {count}")
            
            print(f"\nObject Status:")
            print(f"  Accessible: {status_counts['accessible']}")
            if status_counts['orphaned'] > 0:
                print(f"  Orphaned/Problematic: {status_counts['orphaned']}")
                print(f"  ⚠ Warning: {status_counts['orphaned']} objects have metadata issues")
            
            # Export to Excel
            excel_file = export_objects_to_excel(objects, project_id, project_name)
            if excel_file:
                print(f"\n🎉 Data exported successfully to: {excel_file}")
            else:
                print(f"\n❌ Failed to export data to Excel")
        else:
            print("\n✗ No objects were found or processed")
            
    except Exception as e:
        print(f"✗ Error during object listing: {e}")
    
    finally:
        # Always close connection
        close_zebra_connection(conn)


if __name__ == "__main__":
    main()